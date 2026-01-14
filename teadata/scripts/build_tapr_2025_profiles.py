from __future__ import annotations

import csv
import math
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import-untyped]

RAW_DIR = Path("teadata/data/tapr/raw")
OUT_DIR = Path("teadata/data/tapr")

CAMPUS_STUDENT = RAW_DIR / "2025 Campus Student Information.csv"
CAMPUS_STAFF = RAW_DIR / "2025 Campus Staff Information.csv"
DISTRICT_STUDENT = RAW_DIR / "2025 District Student Information.csv"
DISTRICT_STAFF = RAW_DIR / "2025 District Staff Information.csv"

CAMPUS_OUT = OUT_DIR / "CAMPPROF_2024-2025_FINAL.xlsx"
DISTRICT_OUT = OUT_DIR / "DISTPROF_2024-2025_FINAL.xlsx"
CAMPUS_ELEMENTS_OUT = OUT_DIR / "CAMPPROF_2024-2025_data_elements.xlsx"
DISTRICT_ELEMENTS_OUT = OUT_DIR / "DISTPROF_2024-2025_data_elements.xlsx"


def sanitize_header(s: str) -> str:
    """Convert a column header to a Python-safe identifier."""
    s = str(s)
    s = s.lower()
    s = s.replace(":", "").replace("(", "").replace(")", "")
    s = s.replace("&", " and ")
    s = s.replace("%", " percent ")
    s = s.replace("#", " number ")
    s = s.replace(">", " gt ")
    s = s.replace("<", " lt ")
    s = re.sub(r"[\\/–—\-]+", "_", s)
    s = re.sub(r"[.]", "_", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]+", "_", s)
    s = re.sub(r"__+", "_", s)
    return s.strip("_")


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for header in headers:
        if header not in seen:
            seen[header] = 1
            out.append(header)
            continue
        seen[header] += 1
        out.append(f"{header}_{seen[header]}")
    return out


def _normalize_row(row: list[str], width: int) -> list[str]:
    if len(row) < width:
        row = row + [""] * (width - len(row))
    elif len(row) > width:
        row = row[:width]
    return row


def _read_tapr_csv(
    path: Path, *, rename_map: dict[str, str], key_field: str
):
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        labels = next(reader)
        codes = next(reader)

        headers = [
            rename_map.get(sanitize_header(label), sanitize_header(label))
            for label in labels
        ]
        headers = _dedupe_headers(headers)

        elements = []
        for code, label, header in zip(codes, labels, headers):
            code_text = str(code).strip()
            label_text = str(label).strip()
            elements.append(
                {
                    "code": code_text,
                    "label": label_text,
                    "sanitized_label": header,
                }
            )

        rows: dict[str, dict[str, str]] = {}
        header_len = len(headers)
        for row in reader:
            if not row or not any(cell.strip() for cell in row):
                continue
            row = _normalize_row(row, header_len)
            record = {
                header: (value.strip() if isinstance(value, str) else value)
                for header, value in zip(headers, row)
            }
            key = record.get(key_field)
            if key is None:
                continue
            key = str(key).strip()
            if not key:
                continue
            rows[key] = record

    return headers, rows, elements


def _merge_tables(
    primary_headers: list[str],
    primary_rows: dict[str, dict[str, str]],
    secondary_headers: list[str],
    secondary_rows: dict[str, dict[str, str]],
    *,
    key_field: str,
    preferred_cols: Iterable[str],
):
    keys = sorted(set(primary_rows) | set(secondary_rows))
    merged_rows: list[dict[str, str]] = []
    for key in keys:
        record: dict[str, str] = {}
        record.update(primary_rows.get(key, {}))
        for k, v in secondary_rows.get(key, {}).items():
            if k in record and record[k] not in (None, ""):
                continue
            record[k] = v
        merged_rows.append(record)

    ordered_headers: list[str] = []
    for col in preferred_cols:
        if col in primary_headers or col in secondary_headers:
            if col not in ordered_headers:
                ordered_headers.append(col)

    for col in primary_headers:
        if col not in ordered_headers:
            ordered_headers.append(col)
    for col in secondary_headers:
        if col not in ordered_headers:
            ordered_headers.append(col)

    if key_field not in ordered_headers:
        ordered_headers.insert(0, key_field)

    return ordered_headers, merged_rows


def _rows_for_headers(headers: list[str], records: list[dict[str, str]]) -> list[list[str]]:
    return [[record.get(h, "") for h in headers] for record in records]


_TABLE_STYLE_NAME = "TableStyleMedium16"
_TABLE_NAME_SANITIZER = re.compile(r"[^A-Za-z0-9_]")
_COLUMN_PADDING = 1
_DEFAULT_COLUMN_WIDTH = 12
_MAX_COLUMN_WIDTH = 255
_MIN_CHAR_WIDTH = 4
_MAX_HEADER_LINES = 4
_HEADER_FONT_COLOR = "00FFFFFF"
_HEADER_FILL_COLOR = "FF1F3864"
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrapText=True)


def format_worksheet_as_table(worksheet, *, table_name: str | None = None) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    worksheet.freeze_panes = "A2"
    _wrap_header_row(worksheet)
    _autofit_columns(worksheet)
    _apply_table_style(worksheet, requested_name=table_name)


def _wrap_header_row(worksheet) -> None:
    header_rows = worksheet.iter_rows(min_row=1, max_row=1)
    try:
        header = next(header_rows)
    except StopIteration:
        return

    for cell in header:
        cell.alignment = (cell.alignment or Alignment()).copy(
            horizontal=_HEADER_ALIGNMENT.horizontal,
            vertical=_HEADER_ALIGNMENT.vertical,
            wrapText=_HEADER_ALIGNMENT.wrapText,
        )
        cell.font = (cell.font or Font()).copy(color=_HEADER_FONT_COLOR)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=_HEADER_FILL_COLOR,
            bgColor=_HEADER_FILL_COLOR,
        )


def _autofit_columns(worksheet) -> None:
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_col == 0:
        return

    for column_cells in worksheet.iter_cols(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        header_cell = column_cells[0]
        data_cells = column_cells[1:]

        data_width = max(
            (_cell_display_length(cell.value) for cell in data_cells),
            default=0,
        )
        header_text = _cell_text(header_cell.value)
        header_total_length = _cell_display_length(header_text)
        header_longest_word = _longest_word_length(header_text)

        adjusted_width = data_width + _COLUMN_PADDING if data_width else 0
        adjusted_width = max(adjusted_width, header_longest_word)

        if header_total_length and adjusted_width:
            estimated_lines = math.ceil(header_total_length / adjusted_width)
            if estimated_lines > _MAX_HEADER_LINES:
                widened_for_header = math.ceil(
                    header_total_length / _MAX_HEADER_LINES
                )
                adjusted_width = max(adjusted_width, widened_for_header)

        if adjusted_width == 0:
            adjusted_width = _DEFAULT_COLUMN_WIDTH

        adjusted_width = max(adjusted_width, _MIN_CHAR_WIDTH)
        adjusted_width = min(adjusted_width, _MAX_COLUMN_WIDTH)
        column_letter = get_column_letter(header_cell.column)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _apply_table_style(worksheet, requested_name: str | None) -> None:
    min_row = worksheet.min_row or 1
    max_row = worksheet.max_row or 1
    min_col = worksheet.min_column or 1
    max_col = worksheet.max_column or 1
    if max_row < min_row or max_col < min_col:
        return

    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    table_range = f"{start}:{end}"
    table_name = _generate_table_name(worksheet, requested_name)

    table = Table(displayName=table_name, ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name=_TABLE_STYLE_NAME,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _cell_display_length(value: object) -> int:
    if value in (None, ""):
        return 0
    text = value.decode("utf-8", errors="replace") if isinstance(value, bytes) else str(value)
    if not text:
        return 0
    return max(len(part) for part in text.splitlines()) or 0


def _cell_text(value: object) -> str:
    if value in (None, ""):
        return ""
    return value.decode("utf-8", errors="replace") if isinstance(value, bytes) else str(value)


def _longest_word_length(text: str) -> int:
    return max((len(word) for word in re.split(r"\s+", text) if word), default=0)


def _generate_table_name(worksheet, requested_name: str | None) -> str:
    base_name = requested_name or f"{worksheet.title}_table"
    base_name = _TABLE_NAME_SANITIZER.sub("_", base_name).strip("_") or "Table"
    if not base_name[0].isalpha() and base_name[0] != "_":
        base_name = f"_{base_name}"
    base_name = base_name[:250]

    existing_names = _collect_table_names(worksheet)
    candidate = base_name
    suffix = 1
    while candidate in existing_names:
        candidate = f"{base_name}_{suffix}"
        suffix += 1
    return candidate


def _collect_table_names(worksheet) -> set[str]:
    workbook = getattr(worksheet, "parent", None)
    tables: set[str] = set()
    worksheets = getattr(workbook, "worksheets", []) if workbook is not None else [worksheet]

    for ws in worksheets:
        ws_tables = getattr(ws, "tables", None)
        if not ws_tables:
            continue
        if isinstance(ws_tables, dict):
            tables.update(ws_tables.keys())
        else:
            tables.update(getattr(tbl, "displayName", "") for tbl in ws_tables)

    return {name for name in tables if name}


def _write_workbook(
    output_path: Path,
    *,
    sheet_name: str,
    headers: list[str],
    rows: list[list[str]],
    table_name: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    format_worksheet_as_table(ws, table_name=table_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_elements_workbook(
    output_path: Path,
    *,
    sheet_name: str,
    elements: list[dict[str, str]],
    table_name: str,
) -> None:
    headers = ["code", "label", "sanitized_label", "source"]
    rows = [[el.get(h, "") for h in headers] for el in elements]
    _write_workbook(
        output_path,
        sheet_name=sheet_name,
        headers=headers,
        rows=rows,
        table_name=table_name,
    )


def main() -> None:
    rename_map = {
        "6_digit_county_district_number": "district_number",
        "9_digit_campus_number": "campus_number",
    }

    campus_student_headers, campus_student_rows, campus_student_elements = _read_tapr_csv(
        CAMPUS_STUDENT, rename_map=rename_map, key_field="campus_number"
    )
    campus_staff_headers, campus_staff_rows, campus_staff_elements = _read_tapr_csv(
        CAMPUS_STAFF, rename_map=rename_map, key_field="campus_number"
    )
    district_student_headers, district_student_rows, district_student_elements = _read_tapr_csv(
        DISTRICT_STUDENT, rename_map=rename_map, key_field="district_number"
    )
    district_staff_headers, district_staff_rows, district_staff_elements = _read_tapr_csv(
        DISTRICT_STAFF, rename_map=rename_map, key_field="district_number"
    )

    campus_headers, campus_records = _merge_tables(
        campus_student_headers,
        campus_student_rows,
        campus_staff_headers,
        campus_staff_rows,
        key_field="campus_number",
        preferred_cols=("campus_number", "campus_name", "district_number", "district_name"),
    )
    district_headers, district_records = _merge_tables(
        district_student_headers,
        district_student_rows,
        district_staff_headers,
        district_staff_rows,
        key_field="district_number",
        preferred_cols=("district_number", "district_name"),
    )

    campus_rows = _rows_for_headers(campus_headers, campus_records)
    district_rows = _rows_for_headers(district_headers, district_records)

    _write_workbook(
        CAMPUS_OUT,
        sheet_name="campus_tapr_2025",
        headers=campus_headers,
        rows=campus_rows,
        table_name="campus_tapr_2025",
    )
    _write_workbook(
        DISTRICT_OUT,
        sheet_name="district_tapr_2025",
        headers=district_headers,
        rows=district_rows,
        table_name="district_tapr_2025",
    )

    for el in campus_student_elements:
        el["source"] = "campus_student"
    for el in campus_staff_elements:
        el["source"] = "campus_staff"
    for el in district_student_elements:
        el["source"] = "district_student"
    for el in district_staff_elements:
        el["source"] = "district_staff"

    _write_elements_workbook(
        CAMPUS_ELEMENTS_OUT,
        sheet_name="campus_tapr_elements_2025",
        elements=campus_student_elements + campus_staff_elements,
        table_name="campus_tapr_elements_2025",
    )
    _write_elements_workbook(
        DISTRICT_ELEMENTS_OUT,
        sheet_name="district_tapr_elements_2025",
        elements=district_student_elements + district_staff_elements,
        table_name="district_tapr_elements_2025",
    )


if __name__ == "__main__":
    main()
