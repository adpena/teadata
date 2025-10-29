"""Compare charter campuses against nearby public schools by accountability ratings."""

from __future__ import annotations

import math
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from teadata import DataEngine
from teadata.classes import Campus, District, Query
from teadata.geometry import haversine_miles
from teadata.grades import coerce_grade_spans


RATING_RANK = {"A": 5, "B": 4, "C": 3, "D": 2, "F": 1}
RANK_TO_RATING = {value: key for key, value in RATING_RANK.items()}
DEFAULT_RADIUS_MILES = 3.0
OUTPUT_FILENAME = "charter_vs_public_rating_overlap.xlsx"


@dataclass(frozen=True)
class ComparablePublicCampus:
    campus: Campus
    distance_miles: float
    rating_rank: int


def normalize_rating_label(raw: object) -> Optional[str]:
    if raw is None:
        return None
    if isinstance(raw, float) and math.isnan(raw):
        return None
    value = str(raw).strip()
    if not value or value.lower() == "nan":
        return None
    return value


def rating_to_rank(label: Optional[str]) -> Optional[int]:
    if label is None:
        return None
    normalized = label.strip().upper()
    return RATING_RANK.get(normalized)


def rank_to_rating(rank: Optional[int]) -> Optional[str]:
    if rank is None:
        return None
    return RANK_TO_RATING.get(rank)


def campus_rating_rank(campus: Campus) -> Optional[int]:
    label = normalize_rating_label(
        getattr(campus, "overall_rating_2025", None)
        or getattr(campus, "rating", None)
    )
    return rating_to_rank(label)


def campus_rating_label(campus: Campus) -> Optional[str]:
    return normalize_rating_label(
        getattr(campus, "overall_rating_2025", None)
        or getattr(campus, "rating", None)
    )


def describe_public_campuses(
    comparable: list[ComparablePublicCampus],
) -> dict[str, object]:
    if not comparable:
        return {
            "rating_summary": "",
            "detail_text": "",
            "nearest_name": None,
            "nearest_district": None,
            "nearest_rating": None,
            "best_rank": None,
            "worst_rank": None,
            "average_rank": float("nan"),
            "nearest_distance": float("nan"),
        }

    rating_counts = Counter(rank_to_rating(item.rating_rank) for item in comparable)
    sorted_counts = sorted(
        ((label, count) for label, count in rating_counts.items() if label),
        key=lambda item: -RATING_RANK[item[0]],
    )
    summary_parts = [f"{label}: {count}" for label, count in sorted_counts]
    rating_summary = ", ".join(summary_parts)

    detail_parts: list[str] = []
    sorted_detail = sorted(comparable, key=lambda item: item.distance_miles)
    for item in sorted_detail:
        public = item.campus
        label = rank_to_rating(item.rating_rank)
        detail_parts.append(
            " - ".join(
                [
                    public.name,
                    public.district.name if isinstance(public.district, District) else "Unknown District",
                    public.grade_range or "Unknown Grades",
                    f"Rating {label}",
                    f"{item.distance_miles:.2f} mi",
                ]
            )
        )
    detail_text = "\n".join(detail_parts)

    best_rank = max(item.rating_rank for item in comparable)
    worst_rank = min(item.rating_rank for item in comparable)
    average_rank = sum(item.rating_rank for item in comparable) / len(comparable)

    nearest = sorted_detail[0]
    nearest_label = rank_to_rating(nearest.rating_rank)
    nearest_name = nearest.campus.name
    nearest_district = (
        nearest.campus.district.name if isinstance(nearest.campus.district, District) else None
    )
    return {
        "rating_summary": rating_summary,
        "detail_text": detail_text,
        "nearest_name": nearest_name,
        "nearest_district": nearest_district,
        "nearest_rating": nearest_label,
        "best_rank": best_rank,
        "worst_rank": worst_rank,
        "average_rank": average_rank,
        "nearest_distance": nearest.distance_miles,
    }


def analyze_charter_vs_public(radius_miles: float = DEFAULT_RADIUS_MILES) -> pd.DataFrame:
    engine = DataEngine.from_snapshot(search=True)
    campuses: list[Campus] = [
        campus for campus in engine.campuses.values() if not getattr(campus, "is_private", False)
    ]

    charter_campuses = [campus for campus in campuses if getattr(campus, "is_charter", False)]
    results: list[dict[str, object]] = []

    for charter in charter_campuses:
        charter_coords = getattr(charter, "coords", None)
        if not charter_coords:
            continue
        charter_spans = coerce_grade_spans(charter)
        if not charter_spans:
            continue

        charter_rank = campus_rating_rank(charter)
        if charter_rank is None:
            continue

        nearby = engine.radius_campuses(
            charter_coords[0], charter_coords[1], radius_miles
        )
        grade_spec = getattr(charter, "grade_range", None) or charter

        overlapping = (
            Query(list(nearby), engine) >> ("grade_overlap", grade_spec)
        ).to_list()
        comparable_publics: list[ComparablePublicCampus] = []

        for campus in overlapping:
            if campus.id == charter.id:
                continue
            if getattr(campus, "is_private", False) or getattr(campus, "is_charter", False):
                continue
            public_rank = campus_rating_rank(campus)
            if public_rank is None:
                continue
            public_coords = getattr(campus, "coords", None)
            if not public_coords:
                continue
            distance = haversine_miles(
                charter_coords[0], charter_coords[1], public_coords[0], public_coords[1]
            )
            if distance <= radius_miles:
                comparable_publics.append(
                    ComparablePublicCampus(campus=campus, distance_miles=distance, rating_rank=public_rank)
                )

        if not comparable_publics:
            continue

        min_public_rank = min(item.rating_rank for item in comparable_publics)
        if charter_rank > min_public_rank:
            continue

        public_metrics = describe_public_campuses(comparable_publics)

        result = {
            "Charter Campus": charter.name,
            "Charter Campus Number": charter.campus_number,
            "Charter District": charter.district.name if isinstance(charter.district, District) else None,
            "Charter Grade Range": charter.grade_range,
            "Charter Rating": campus_rating_label(charter),
            "Comparable Public Campuses": len(comparable_publics),
            "Nearest Public Campus": public_metrics["nearest_name"],
            "Nearest Public District": public_metrics["nearest_district"],
            "Nearest Public Rating": public_metrics["nearest_rating"],
            "Nearest Public Distance (mi)": round(public_metrics["nearest_distance"], 2)
            if isinstance(public_metrics["nearest_distance"], (int, float))
            else public_metrics["nearest_distance"],
            "Best Public Rating": rank_to_rating(public_metrics["best_rank"])
            if public_metrics["best_rank"] is not None
            else None,
            "Worst Public Rating": rank_to_rating(public_metrics["worst_rank"])
            if public_metrics["worst_rank"] is not None
            else None,
            "Average Public Rating Score": round(public_metrics["average_rank"], 2)
            if isinstance(public_metrics["average_rank"], (int, float))
            else public_metrics["average_rank"],
            "Public Rating Summary": public_metrics["rating_summary"],
            "Public Campus Details": public_metrics["detail_text"],
        }
        results.append(result)

    df = pd.DataFrame(results)
    if df.empty:
        return df

    sort_cols = ["Charter Rating", "Charter District", "Charter Campus"]
    df = df.sort_values(sort_cols).reset_index(drop=True)
    return df


def style_excel(output_path: Path, df: pd.DataFrame) -> None:
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        print("openpyxl styling unavailable; exported raw workbook.")
        return

    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            value_length = len(str(value))
            if value_length > max_length:
                max_length = value_length
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    detail_col_index = df.columns.get_loc("Public Campus Details") + 1
    detail_column_letter = get_column_letter(detail_col_index)
    for cell in ws[detail_column_letter]:
        cell.alignment = wrap_alignment

    for column_name in [
        "Charter Rating",
        "Nearest Public Rating",
        "Best Public Rating",
        "Worst Public Rating",
    ]:
        idx = df.columns.get_loc(column_name) + 1
        letter = get_column_letter(idx)
        for cell in ws[letter]:
            cell.alignment = center_alignment

    wb.save(output_path)


def export_to_excel(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Charter vs Public")
    style_excel(output_path, df)


def main() -> None:
    df = analyze_charter_vs_public()
    if df.empty:
        print("No charter campuses met the comparison criteria.")
        return
    output_path = Path(__file__).with_name(OUTPUT_FILENAME)
    export_to_excel(df, output_path)
    print(
        f"Identified {len(df)} charter campuses meeting the criteria. Exported results to {output_path}"
    )


if __name__ == "__main__":
    main()
